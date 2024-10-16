import os
import requests
from datetime import datetime
from dotenv import load_dotenv
import json
import sqlite3
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from tqdm import tqdm

# Load environment variables from .env file
load_dotenv()

def get_access_token():
    url = "https://api.lookout.com/oauth2/token"
    headers = {
        "Content-Type": "application/x-www-form-urlencoded",
        "Authorization": f"Bearer {os.environ['REACT_APP_APPLICATION_KEY']}"
    }
    data = {"grant_type": "client_credentials"}
    response = requests.post(url, headers=headers, data=data)
    return response.json()["access_token"]

def get_devices(access_token, email=None):
    url = "https://api.lookout.com/mra/api/v2/devices"
    headers = {
        "Authorization": f"Bearer {access_token}",
        "Accept": "application/json"
    }
    params = {
        "limit": 1000
    }
    if email:
        params["email"] = email
    
    devices = []
    total_devices = 0
    with tqdm(desc="Fetching devices", unit="device") as pbar:
        while True:
            response = requests.get(url, headers=headers, params=params)
            data = response.json()
            batch = data["devices"]
            devices.extend(batch)
            total_devices += len(batch)
            pbar.update(len(batch))
            pbar.set_postfix({"Total": total_devices})
            
            if data["count"] < params["limit"]:
                break
            elif "oid" in batch[-1]:
                params["oid"] = batch[-1]["oid"]
            else:
                print("Warning: Unable to paginate further. Some devices may be missing.")
                break
    
    print(f"Total devices fetched: {total_devices}")
    return devices

def get_threats(access_token):
    url = "https://api.lookout.com/mra/api/v2/threats"
    headers = {
        "Authorization": f"Bearer {access_token}",
        "Accept": "application/json"
    }
    params = {
        "limit": 1000
    }
    
    threats = []
    total_threats = 0
    with tqdm(desc="Fetching threats", unit="threat") as pbar:
        while True:
            response = requests.get(url, headers=headers, params=params)
            data = response.json()
            batch = data["threats"]
            threats.extend(batch)
            total_threats += len(batch)
            pbar.update(len(batch))
            pbar.set_postfix({"Total": total_threats})
            
            if data["count"] < params["limit"]:
                break
            elif "oid" in batch[-1]:
                params["oid"] = batch[-1]["oid"]
            else:
                print("Warning: Unable to paginate further. Some threats may be missing.")
                break
    
    print(f"Total threats fetched: {total_threats}")
    return threats

def calculate_threat_age(detected_at):
    formats = ['%Y-%m-%dT%H:%M:%S.%fZ', '%Y-%m-%dT%H:%M:%S']
    for fmt in formats:
        try:
            detected_time = datetime.strptime(detected_at, fmt)
            age = datetime.utcnow() - detected_time
            return age.days
        except ValueError:
            continue
    raise ValueError(f"Unable to parse date: {detected_at}")

def create_database():
    conn = sqlite3.connect('devices.db')
    c = conn.cursor()
    
    c.execute('''CREATE TABLE IF NOT EXISTS devices
                 (guid TEXT PRIMARY KEY, email TEXT, checkin_time TEXT, 
                 os_version TEXT, protection_status TEXT, 
                 manufacturer TEXT, model TEXT, 
                 latest_os_version TEXT, latest_security_patch_level TEXT, 
                 security_patch_level TEXT, sdk_version TEXT, platform TEXT)''')
    
    conn.commit()
    conn.close()

def update_device_database(devices):
    conn = sqlite3.connect('devices.db')
    c = conn.cursor()
    with tqdm(total=len(devices), desc="Updating database", unit="device") as pbar:
        for device in devices:
            c.execute('''INSERT OR REPLACE INTO devices 
                         (guid, email, checkin_time, os_version, protection_status, 
                         manufacturer, model, latest_os_version, latest_security_patch_level, 
                         security_patch_level, sdk_version, platform)
                         VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)''',
                      (device.get('guid', 'N/A'),
                       device.get('email', 'N/A'),
                       device.get('checkin_time', 'N/A'),
                       device.get('software', {}).get('os_version', 'N/A'),
                       device.get('protection_status', 'N/A'),
                       device.get('hardware', {}).get('manufacturer', 'N/A'),
                       device.get('hardware', {}).get('model', 'N/A'),
                       device.get('software', {}).get('latest_os_version', 'N/A'),
                       device.get('software', {}).get('latest_security_patch_level', 'N/A'),
                       device.get('software', {}).get('security_patch_level', 'N/A'),
                       device.get('software', {}).get('sdk_version', 'N/A'),
                       device.get('platform', 'N/A')))
            pbar.update(1)
    conn.commit()
    conn.close()

def get_device_info(guid):
    conn = sqlite3.connect('devices.db')
    c = conn.cursor()
    c.execute('SELECT * FROM devices WHERE guid = ?', (guid,))
    device = c.fetchone()
    conn.close()
    if device:
        return {
            'guid': device[0],
            'email': device[1],
            'checkin_time': device[2],
            'os_version': device[3],
            'protection_status': device[4],
            'manufacturer': device[5],
            'model': device[6],
            'latest_os_version': device[7],
            'latest_security_patch_level': device[8],
            'security_patch_level': device[9],
            'sdk_version': device[10],
            'platform': device[11]
        }
    return None

def refresh_device_data(access_token):
    devices = get_devices(access_token)
    update_device_database(devices)
    print(f"Device database refreshed with {len(devices)} devices.")

def generate_excel_report(device_info, threat_age_buckets):
    wb = Workbook()
    ws = wb.active
    ws.title = "Device and Threat Report"

    # Define styles
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    wrap_text = Alignment(wrap_text=True)

    # Write headers
    headers = ["Device GUID", "User Email", "Platform", "Manufacturer", "Model", "OS Version", "Latest OS Version",
               "Security Patch Level", "Latest Security Patch Level", "SDK Version", "Last Check-in",
               "Connection Status", "Threats"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill

    # Write device data
    print("Generating Excel report...")
    with tqdm(total=len(device_info), desc="Writing device data", unit="device") as pbar:
        row = 2
        for guid, info in device_info.items():
            ws.cell(row=row, column=1, value=guid)
            ws.cell(row=row, column=2, value=info.get('email', 'N/A'))
            ws.cell(row=row, column=3, value=info.get('platform', 'N/A'))
            ws.cell(row=row, column=4, value=info.get('manufacturer', 'N/A'))
            ws.cell(row=row, column=5, value=info.get('model', 'N/A'))
            ws.cell(row=row, column=6, value=info.get('os_version', 'N/A'))
            ws.cell(row=row, column=7, value=info.get('latest_os_version', 'N/A'))
            ws.cell(row=row, column=8, value=info.get('security_patch_level', 'N/A'))
            ws.cell(row=row, column=9, value=info.get('latest_security_patch_level', 'N/A'))
            ws.cell(row=row, column=10, value=info.get('sdk_version', 'N/A'))
            ws.cell(row=row, column=11, value=info.get('checkin_time', 'N/A'))
            ws.cell(row=row, column=12, value='Disconnected' if info.get('is_disconnected', False) else 'Connected')

            # Combine threat information into a single cell
            threats = info.get('threats', [])
            if threats:
                threat_info = "\n".join([f"Name: {t['threat_name']}, Age: {t['age_days']} days, Status: {t['status']}, Risk: {t['risk']}" for t in threats])
            else:
                threat_info = "No threats detected"
            cell = ws.cell(row=row, column=13, value=threat_info)
            cell.alignment = wrap_text

            row += 1
            pbar.update(1)

    print("Adjusting column widths...")
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column].width = adjusted_width

    # Add a new sheet for threat age buckets
    ws_threats = wb.create_sheet(title="Threat Aging")
    ws_threats.append(["Age Bucket", "Number of Threats"])
    for bucket, count in threat_age_buckets.items():
        ws_threats.append([bucket, count])

    # Save the workbook
    print("Saving Excel report...")
    wb.save("device_and_threat_report.xlsx")
    print("Excel report generated: device_and_threat_report.xlsx")

if __name__ == "__main__":
    print("Starting the Device and Threat Reporting Tool...")
    access_token = get_access_token()
    print("Access token obtained.")
    
    create_database()
    print("Database created/verified.")
    
    refresh_device_data(access_token)
    
    print("Fetching threat data...")
    threats = get_threats(access_token)
    
    # Create a dictionary to store all device information
    device_info = defaultdict(lambda: {"threats": [], "is_disconnected": False})

    # Process threats and get device info from database
    threat_age_buckets = {'< 1 day': 0, '1-7 days': 0, '8-30 days': 0, '31-90 days': 0, '> 90 days': 0}
    print("Processing threats and device information...")
    with tqdm(total=len(threats), desc="Processing threats", unit="threat") as pbar:
        for threat in threats:
            device_guid = threat.get('device_guid', 'N/A')
            db_device_info = get_device_info(device_guid)
            if db_device_info:
                device_info[device_guid].update(db_device_info)
                device_info[device_guid]['is_disconnected'] = db_device_info['protection_status'] == 'DISCONNECTED'
            
            detected_at = threat.get('detected_at')
            if detected_at:
                try:
                    age_days = calculate_threat_age(detected_at)
                    threat_info = {
                        'threat_name': threat.get('classification', 'Unknown'),
                        'age_days': age_days,
                        'status': threat.get('status', 'Unknown'),
                        'risk': threat.get('risk', 'Unknown')
                    }
                    device_info[device_guid]['threats'].append(threat_info)
                    
                    if threat.get('status') != 'RESOLVED':
                        if age_days < 1:
                            threat_age_buckets['< 1 day'] += 1
                        elif age_days < 8:
                            threat_age_buckets['1-7 days'] += 1
                        elif age_days < 31:
                            threat_age_buckets['8-30 days'] += 1
                        elif age_days < 91:
                            threat_age_buckets['31-90 days'] += 1
                        else:
                            threat_age_buckets['> 90 days'] += 1
                except ValueError as e:
                    print(f"Error processing threat: {e}")
            pbar.update(1)

    # Generate Excel report
    generate_excel_report(device_info, threat_age_buckets)
    
    print("Script execution completed.")
